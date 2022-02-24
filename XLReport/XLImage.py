# define image for excel files
import collections
import openpyxl as xl
from openpyxl.drawing.image import Image as xlImage
from XLReport.XLBase import DrawUnderLine, SetColWidth, RowCol_toCellStr, SetRowHeight_inPixel
from XLReport.XLBase import Styles as BaseStyles

Styles = {
    'Font': BaseStyles['Font']['Italic'],
    'Align': BaseStyles['Align']['Left'],
    'NameRowLine': BaseStyles['Border']['ThinUnder'],
    'ImageGuideLine': BaseStyles['Border']['DoubleLeft'],
    'WidthMargin': 0.5
}

ImageIDKey = "__FelineReportImage"
RowOccupied = 3
ColOccupied = 2


def IsImage(jsonObj):
    return isinstance(jsonObj, Image) or _isImage_json(jsonObj)


def _isImage_json(jsonObj):
    return isinstance(jsonObj, collections.Mapping) and \
        (len(jsonObj.keys()) == 1) and (ImageIDKey in jsonObj)


def CountColImage(obj):
    return ColOccupied


def IndentSizeImage(obj):
    return BaseStyles['RowSize']['Indent'] * 2


def TryImageConversion(obj):
    if _isImage_json(obj):
        obj = Image(obj[ImageIDKey])
        return True
    elif isinstance(obj, Image):
        return True
    else:
        return False


def InsertImage(worksheet, baseRow, baseCol, valueCol, name, obj):
    if TryImageConversion(obj):
        return obj.Insert(worksheet, baseRow, baseCol, valueCol, name, autofit=True)
    else:
        return False


class Image:
    def __init__(self, path):
        self.Path = path

    def _computeImageWidth(self, worksheet, indentCnt):
        # unit of the margins are in character width 11
        pageWidth = BaseStyles['RowSize']['PageWidth']
        indentsize = 0
        for col in range(1, indentCnt):
            indentsize += worksheet.column_dimensions[
                xl.utils.get_column_letter(col)].width
        return (pageWidth - indentsize - 2) * worksheet.sheet_format.baseColWidth

    # returns image size
    def _setImageSize_bySize(self, img, width, height):
        img.width = width
        img.height = height
        return [width, height]

    # returns image size
    def _setImageSize_byRatio(self, img, ratio):
        img.width *= ratio
        img.height *= ratio
        return [img.width, img.height]

    # returns image size
    def _setImageSize_autoFit(self, img, worksheet, indentCnt):
        imgWidth = self._computeImageWidth(worksheet, indentCnt)
        scaleRatio = imgWidth / img.width
        return self._setImageSize_byRatio(img, scaleRatio)

    def Insert(self, worksheet, baseRow, baseCol, valueCol, name, width=None, height=None, ratio=None, autofit=True):
        # write current obj info
        # set value/style for name cell
        nameC = worksheet.cell(row=baseRow, column=baseCol, value=name)
        nameC.font = Styles['Font']
        nameC.alignment = Styles['Align']
        # draw undeline under name row
        DrawUnderLine(worksheet, baseRow, baseCol,
                      valueCol, Styles['NameRowLine'])
        # # set indent size
        # SetColWidth(worksheet, baseCol,
        #             BaseStyles['RowSize']['Indent'])
        # SetColWidth(worksheet, baseCol+1,
        #             BaseStyles['RowSize']['Indent'])
        # merges cells to put in the image
        worksheet.merge_cells(
            start_row=baseRow+1, start_column=valueCol-1, end_row=baseRow+1, end_column=valueCol)
        # draw guide line
        worksheet.cell(row=baseRow+1, column=baseCol +
                       1).border = Styles['ImageGuideLine']
        # import/adjust size/insert image
        img = xlImage(self.Path)
        # resize image autofit:
        if autofit:
            self._setImageSize_autoFit(img, worksheet, valueCol-1)
        elif ratio is not None:
            self._setImageSize_byRatio(img, ratio=ratio)
        elif (width is not None) and (height is not None):
            self._setImageSize_bySize(img, width=width, height=height)
        else:
            self._setImageSize_autoFit(img, worksheet, valueCol-1)
        # insert image at baseRow+1, baseCol+2
        worksheet.add_image(img, RowCol_toCellStr(baseRow+1, valueCol-1))
        # adjust row height
        SetRowHeight_inPixel(worksheet, baseRow+1, img.height)
        # return occupying rows
        return ColOccupied
