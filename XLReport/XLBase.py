import openpyxl as xl

# define styles for the report sheets
Styles = {
    "Font": {
        'Base': xl.styles.Font(size=7),
        'Bold': xl.styles.Font(size=7, b=True),
        'Italic': xl.styles.Font(size=7, i=True),
        'Title': xl.styles.Font(size=9, b=True),
        'Data': xl.styles.Font(size=7.5),
        'DataBold': xl.styles.Font(size=7.5, b=True)
    },
    "Align": {
        'Left': xl.styles.Alignment(horizontal="left", vertical="center"),
        'Right': xl.styles.Alignment(horizontal="right", vertical="center"),
        'Center': xl.styles.Alignment(horizontal="center", vertical="center"),
        'LeftWithWrap': xl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True),
    },
    "Border": {
        'ThinUnder':  xl.styles.Border(bottom=xl.styles.Side('thin')),
        'ThinRight':  xl.styles.Border(right=xl.styles.Side('thin')),
        'ThinAllround': xl.styles.Border(
            top=xl.styles.Side('thin'),
            bottom=xl.styles.Side('thin'),
            right=xl.styles.Side('thin'),
            left=xl.styles.Side('thin')),
        'ThickUnder':  xl.styles.Border(bottom=xl.styles.Side('thick')),
        'DoubleUnder':  xl.styles.Border(bottom=xl.styles.Side('double')),
        'DoubleLeft': xl.styles.Border(left=xl.styles.Side('double')),
    },
    "Fill": {
        'LightShade': xl.styles.PatternFill('solid', fgColor='e0e0e0')
    },
    "RowSize": {
        "Indent": 2,
        "ArrayIndex": 7,
        "ValueCell": 20,
        "PageWidth": 57
    },
    # set papersize to B5 (B5code = 13)
    'PageSetup': xl.worksheet.page.PrintPageSetup(paperSize=13),
    'PageMargins': xl.worksheet.page.PageMargins(left=1.2, right=1.2, top=0.8, bottom=0.6, header=0.25, footer=0.25),
    'PrintOpt': xl.worksheet.page.PrintOptions(horizontalCentered=True, verticalCentered=False),
    'View': 'pageLayout',
    'DatatableMargin': 2
}


def DrawUnderLine(worksheet, row, startCol, endCol, lineType):
    for col in range(startCol, endCol + 1):
        worksheet.cell(row=row, column=col).border = lineType


def RowCol_toCellStr(row, col):
    return xl.utils.get_column_letter(col) + str(row)


def SetRowHeight_inPixel(worksheet, row, heightInPixel):
    worksheet.row_dimensions[row].height = \
        xl.utils.units.pixels_to_points(heightInPixel)


def SetRowHeight(worksheet, row, heightInPoint):
    worksheet.row_dimensions[row].height = heightInPoint


def SetColWidth(worksheet, col, widthInCharUnit):
    worksheet.column_dimensions[
        xl.utils.get_column_letter(col)].width = widthInCharUnit


def SetPrintArea(worksheet, startRow, startCol, endRow, endCol):
    startCellStr = RowCol_toCellStr(startRow, startCol)
    endCellStr = RowCol_toCellStr(endRow, endCol)
    worksheet.print_area = startCellStr + ':' + endCellStr


def SetHeaderFooter(worksheet, titleStr):
    if titleStr is not None:
        worksheet.oddHeader.center.text = titleStr + ": &[Tab]"
    else:
        worksheet.oddHeader.center.text = "&[Tab]"
    worksheet.oddHeader.center.size = 9
    worksheet.oddHeader.center.font = "Bold"
    worksheet.oddHeader.center.color = "D0D0D0"
    worksheet.oddFooter.center.text = "&[Page] / &[Pages]"
    worksheet.oddFooter.center.size = 7
    worksheet.oddFooter.center.color = "D0D0D0"
