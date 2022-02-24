import collections
import statistics
import copy
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.worksheet import worksheet
import XLReport


def BuildSheet_Summary(ws, report):
    # get needed data
    pageData = {
        'Experiment name': report['ExperimentName'],
        'Experiment comments': report['ExperimentComments'],
        'Experiment date': report['ExperimentDatetime'],
        'Input parameters': report['InputParams'],
        'Input trajectories': {
            'Number of clusters': len(report['IndexList']),
            'Input trajectories': XLReport.Image(report['Images']['InputTrajs'])
        },
        'Clusetring results': {
            "Number of clusters": len(report['ClusteringResults']),
            "Clustered trajectories (without noise)": XLReport.Image(report['Images']['ClusteringResult'])
        }
    }

    XLReport.BuildSheet(ws, pageData,
                        "Experiment Summary",
                        report['ExperimentName'])


def BuildSheet_Inputs(ws, report):
    # get length of item(to compute column width of traj#)
    dataCnt = len(report['IndexList'])
    # set page data
    pageData = {
        "Input trajectories": XLReport.Datatable(
            data=report['IndexList'], insertIndex=True,
            header=['Trajectory file path', 'Traj#'], insertHeader=True,
            colWidthFormat=[82, len(str(dataCnt)) + 2])
    }
    # build sheet
    XLReport.BuildSheet(ws, pageData, noPrint=True)
    # insert traj img
    imgpath = report['Images']['InputTrajs']
    img = xl.drawing.image.Image(imgpath)
    ws.add_image(img, 'D3')


def BuildSheet_DistanceMatrix(ws, report):
    pageData = {
        "Distance matrix": XLReport.Datatable(
            data=report['DistanceMatrix'],
            insertIndex=True, insertHeader=True,
            colWidthFormat=6)
    }
    XLReport.BuildSheet(ws, pageData, noPrint=True)
    # get needed data
    distMatrix = report['DistanceMatrix']


def _min(arr):
    minval = float('inf')
    for item in arr:
        if XLReport.IsArray(item):
            item_minval = _min(item)
        else:
            item_minval = item

        if minval > item_minval:
            minval = item_minval
    return minval


def _max(arr):
    maxval = float('-inf')
    for item in arr:
        if XLReport.IsArray(item):
            item_maxval = _max(item)
        else:
            item_maxval = item

        if maxval < item_maxval:
            maxval = item_maxval
    return maxval


def _sum_withCnt(arr):
    sum = 0
    cnt = 0
    for item in arr:
        if XLReport.IsArray(item):
            internal = _sum_withCnt(item)
            sum += internal[0]
            cnt += internal[1]
        else:
            sum += item
            cnt += 1
    return [sum, cnt]


def _mean(arr):
    sum = 0
    cnt = 0
    for item in arr:
        if XLReport.IsArray(item):
            internal = _sum_withCnt(item)
            sum += internal[0]
            cnt += internal[1]
        else:
            sum += item
            cnt += 1
    return sum / cnt


def _sqrdSum_withCnt(arr, mean):
    sqrdSum = 0
    cnt = 0
    for item in arr:
        if XLReport.IsArray(item):
            internal = _sqrdSum_withCnt(item, mean)
            sqrdSum += internal[0]
            cnt += internal[1]
        else:
            sqrdSum += ((item - mean)*(item - mean))
            cnt += 1
    return [sqrdSum, cnt]


def _stddev(arr):
    mean = _mean(arr)
    sqrdSum = 0
    cnt = 0
    for item in arr:
        if XLReport.IsArray(item):
            internal = _sqrdSum_withCnt(item, mean)
            sqrdSum += internal[0]
            cnt += internal[1]
        else:
            sqrdSum += ((item - mean)*(item - mean))
            cnt += 1
    return sqrdSum / cnt


# build sheet Result Summary
def BuildSheet_ResultSummary(ws, report):
    # build page data from report
    HaveNoiseCluster = report['IsNoiseClusterDefined']
    pageData = {
        'Experiment name': report['ExperimentName'],
        'Experiment date': report['ExperimentDatetime'],
        'Total number of clusters': len(report['ClusteringResults']),
        'Clustered trajectories (w/o noise)': XLReport.Image(report['Images']['ClusteringResult'])
    }
    if HaveNoiseCluster:
        pageData['Clustered trajectories (with noise)'] = \
            XLReport.Image(report['Images']['ClusteringResult_withNoise'])
    pageData['Representative routes'] = XLReport.Image(
        report['Images']['Representatives'])
    if HaveNoiseCluster:
        pageData['Result clusters'] = {}
        pageData['Result clusters']['Clusters'] = report['ClusteringResults'][1:]
        pageData['Result clusters']['Noise trajectories'] = \
            [report['ClusteringResults'][0]]
    else:
        pageData['Result clusters'] = report['ClusteringResults']

    pageData['Evaluation'] = {}
    if "RandIndex" in report['Evaluation']:
        pageData['Evaluation']['Rand index'] = report['Evaluation']['RandIndex']['Total']
    if "ClassificationError" in report['Evaluation']:
        pageData['Evaluation']['Classification error'] = report['Evaluation']['ClassificationError']
    if "VariationOfInformation" in report['Evaluation']:
        pageData['Evaluation']['Variation of information (VI)'] = report['Evaluation'][
            'VariationOfInformation']
    pageData['Evaluation']['Davies-Bouldin index'] = report['Evaluation']['DaviesBouldin']
    # compute statistics for Silhouettes. exclude noise cluster
    silhouettesNP = copy.deepcopy(report['Evaluation']['Silhouettes'])
    if report['IsNoiseClusterDefined']:
        if "Noise" in report['ClusterNames']:
            silhouettesNP.pop(report['ClusterNames'].index("Noise"))

    pageData['Evaluation']['Silhouettes'] = {
        'Min value': _min(silhouettesNP),
        'Max value': _max(silhouettesNP),
        'Mean value': _mean(silhouettesNP),
        'Standard deviation': _stddev(silhouettesNP),
        'Silhouettes': XLReport.Image(report['Images']['Silhouettes'])
    }
    XLReport.BuildSheet(worksheet=ws, report=pageData,
                        title='Experiment Results', headerTitle=report['ExperimentName'])


def BuildSheet_ResultCluster(ws, report, idx):
    # max string length for trajectory number (datatable)
    datatableTrajNumCol = len(str(len(report['IndexList']))) + 2
    # build page data from report
    pageData = {
        'Cluster name': report['ClusterNames'][idx],
        'Number of trajectories': len(report['ClusteringResults'][idx]),
        'Clustered trajectories': XLReport.Image(report['Images']['Clusters'][idx]),
        'Cluster trajectory indices': XLReport.Datatable(
            data=[report['IndexList'][i]
                  for i in report['ClusteringResults'][idx]],
            index=report['ClusteringResults'][idx],
            header=['Trajectory file path', 'Traj#'],
            colWidthFormat=[80, datatableTrajNumCol]),
        'Evaluation': {}
    }
    if "RandIndex" in report['Evaluation']:
        pageData['Evaluation']['Rand index'] = report['Evaluation']['RandIndex']['byCluster'][idx]
    pageData['Evaluation']['Davies-Bouldin index'] = report['Evaluation']['DaviesBouldin'][idx]
    pageData['Evaluation']['Silhouettes'] = {
        'Min value': min(report['Evaluation']['Silhouettes'][idx]),
        'Max value': max(report['Evaluation']['Silhouettes'][idx]),
        'Mean value': statistics.mean(report['Evaluation']['Silhouettes'][idx])
    }
    if len(report['Evaluation']['Silhouettes'][idx]) > 2:
        pageData['Evaluation']['Silhouettes']['Standard deviation'] = \
            statistics.stdev(report['Evaluation']['Silhouettes'][idx])
    pageData['Evaluation']['Silhouettes']['Silhouettes'] = \
        XLReport.Image(report['Images']['SilhouettesCluster'][idx])
    pageData['Evaluation']['Silhouettes']['Values'] = \
        XLReport.Datatable(data=report['Evaluation']['Silhouettes'][idx],
                           insertIndex=True, insertHeader=False, colWidthFormat=10)
    sheetTitle = "Experiment Results(" + report['ClusterNames'][idx] + ")"
    XLReport.BuildSheet(ws, pageData,
                        title=sheetTitle,
                        headerTitle=report['ExperimentName'])


def FelineReportXL(jsonData, rootDir):
    # create a workbook(new xl file)
    wb = Workbook()
    # create & worksheets in xl file
    # Summary
    wb.active.title = 'Summary'  # rename default worksheet for first page
    BuildSheet_Summary(wb['Summary'], jsonData)

    # Inputs
    wb.create_sheet('Inputs')
    BuildSheet_Inputs(wb['Inputs'], jsonData)

    # distance matrix
    wb.create_sheet('Distance Matrix')
    BuildSheet_DistanceMatrix(wb['Distance Matrix'], jsonData)

    # result summary
    wb.create_sheet('Result Summary')
    BuildSheet_ResultSummary(wb['Result Summary'], jsonData)

    # result clusters
    for sheetIdx, name in enumerate(jsonData['ClusterNames']):
        sheetname = 'Result ' + name
        wb.create_sheet(sheetname)
        BuildSheet_ResultCluster(wb[sheetname], jsonData, sheetIdx)

    # write to excel file
    XLFilePath = jsonData['InputParams']['outputDir'] \
        + '/' + jsonData['ExperimentName'] + '/Report/' \
        + '/' + jsonData['ExperimentName'] + '_Report.xlsx'
    # save to file
    wb.save(XLFilePath)
